#!/usr/bin/env python3
"""
Final Listesi Birleştirme Aracı
Excel dosyalarını birleştir
"""

import customtkinter as ctk
from tkinter import filedialog, messagebox, ttk
from pathlib import Path
import pandas as pd
import threading
import subprocess
import sys
import shutil
from datetime import datetime
import os
import json
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
import time

try:
    from tkinterdnd2 import DND_FILES, TkinterDnD
    HAS_DND = True
except ImportError:
    HAS_DND = False

ctk.set_appearance_mode("light")
ctk.set_default_color_theme("blue")

CURRENCY_SYMBOLS = {
    'EUR': '€', 'USD': '$', 'GBP': '£', 'TRY': '₺', 'JPY': '¥', 'CNY': '¥',
}


def _get_script_dir():
    if getattr(sys, 'frozen', False):
        return Path(sys.executable).parent
    return Path(__file__).parent


SETTINGS_FILE = _get_script_dir() / '.merger_settings.json'


class Tooltip:
    """Widget üzerine gelince açıklama baloncuğu gösterir"""
    def __init__(self, widget, text):
        self.widget = widget
        self.text = text
        self.tip_window = None
        self._after_id = None
        widget.bind('<Enter>', self._schedule_show)
        widget.bind('<Leave>', self._hide)

    def _schedule_show(self, event=None):
        self._cancel()
        self._after_id = self.widget.after(400, self._show)

    def _show(self):
        self._after_id = None
        if self.tip_window:
            return
        import tkinter as tk
        x = self.widget.winfo_rootx() + 20
        y = self.widget.winfo_rooty() + self.widget.winfo_height() + 5
        self.tip_window = tw = tk.Toplevel(self.widget)
        tw.wm_overrideredirect(True)
        tw.wm_geometry(f"+{x}+{y}")
        tw.attributes('-topmost', True)
        tw.attributes('-disabled', True)
        label = tk.Label(
            tw, text=self.text,
            font=("Segoe UI", 11),
            bg="#34495E", fg="white",
            padx=10, pady=5
        )
        label.pack()

    def _hide(self, event=None):
        self._cancel()
        if self.tip_window:
            self.tip_window.destroy()
            self.tip_window = None

    def _cancel(self):
        if self._after_id:
            self.widget.after_cancel(self._after_id)
            self._after_id = None


class FinalListMerger:
    def __init__(self, root):
        self.root = root
        self.uploaded_files = []
        self.file_item_counts = {}
        self.template_path = None
        self.output_path = None
        self.custom_output_dir = None
        self.is_processing = False
        self._pulsing = False
        self._all_buttons = []

        self._last_browse_dir = self._load_setting('last_browse_dir', '')

        # Openpyxl stil objeleri (her satırda yeniden oluşturmamak için)
        self._thin_border = Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'), bottom=Side(style='thin')
        )
        self._no_border = Border()
        self._header_fill = PatternFill(start_color='3498DB', end_color='3498DB', fill_type='solid')
        self._header_font = Font(bold=True, size=11, color='FFFFFF')
        self._center_align = Alignment(horizontal='center', vertical='center')
        self._data_align = Alignment(vertical='center', wrap_text=True)
        self._bold_font = Font(bold=True, size=11)
        self._right_align = Alignment(horizontal='right', vertical='center')

        self.setup_ui()
        self._setup_dnd()

    # ── Ayarlar ──────────────────────────────────────────────

    def _load_setting(self, key, default=None):
        try:
            if SETTINGS_FILE.exists():
                with open(SETTINGS_FILE, 'r', encoding='utf-8') as f:
                    return json.load(f).get(key, default)
        except Exception:
            pass
        return default

    def _save_setting(self, key, value):
        try:
            settings = {}
            if SETTINGS_FILE.exists():
                with open(SETTINGS_FILE, 'r', encoding='utf-8') as f:
                    settings = json.load(f)
            settings[key] = value
            with open(SETTINGS_FILE, 'w', encoding='utf-8') as f:
                json.dump(settings, f)
        except Exception:
            pass

    # ── Drag & Drop ──────────────────────────────────────────

    def _setup_dnd(self):
        if not HAS_DND:
            return
        try:
            self.root.drop_target_register(DND_FILES)
            self.root.dnd_bind('<<Drop>>', self._on_drop)
            self.drop_area.configure(
                text="📂 Dosya Sürükle & Bırak\nveya tıkla"
            )
        except Exception:
            pass

    def _on_drop(self, event):
        files = self.root.tk.splitlist(event.data)
        added = False
        for file_path in files:
            path = Path(file_path)
            if path.suffix.lower() == '.xlsx' and path not in self.uploaded_files:
                self.uploaded_files.append(path)
                added = True
        if added:
            self._scan_and_update()

    # ── UI ───────────────────────────────────────────────────

    def setup_ui(self):
        self.root.title("Final Listesi Birleştirme Aracı")
        self.root.geometry("900x830")
        self.root.minsize(700, 650)

        self.root.grid_rowconfigure(0, weight=1)
        self.root.grid_columnconfigure(0, weight=1)

        main_container = ctk.CTkScrollableFrame(self.root, fg_color="#F8F9FA")
        main_container.grid(row=0, column=0, sticky="nsew")
        main_container.grid_columnconfigure(0, weight=1)

        # HEADER
        header_frame = ctk.CTkFrame(main_container, fg_color="#FFFFFF")
        header_frame.grid(row=0, column=0, sticky="ew", padx=0, pady=0)
        header_frame.grid_columnconfigure(0, weight=1)

        title_container = ctk.CTkFrame(header_frame, fg_color="#FFFFFF")
        title_container.grid(row=0, column=0, sticky="ew", padx=30, pady=20)
        title_container.grid_columnconfigure(1, weight=1)

        ctk.CTkLabel(title_container, text="📋", font=("Segoe UI", 32, "bold"), text_color="#2C3E50").grid(row=0, column=0, padx=(0, 15))
        ctk.CTkLabel(title_container, text="Final Listesi Birleştirme Aracı", font=("Segoe UI", 28, "bold"), text_color="#2C3E50").grid(row=0, column=1, sticky="w")
        ctk.CTkLabel(title_container, text="Excel dosyalarını birleştir", font=("Segoe UI", 12), text_color="#7F8C8D").grid(row=1, column=0, columnspan=2, sticky="w", pady=(5, 0))

        # CONTENT
        content_frame = ctk.CTkFrame(main_container, fg_color="#F8F9FA")
        content_frame.grid(row=1, column=0, sticky="nsew", padx=20, pady=20)
        content_frame.grid_columnconfigure(0, weight=1)

        # ── UPLOAD CARD ──
        upload_card = self._create_card(content_frame, "📂 Dosya Seçimi")
        upload_card.grid(row=0, column=0, sticky="ew", pady=(0, 20))

        self.drop_area = ctk.CTkButton(
            upload_card,
            text="👆 Dosya Seçmek İçin Tıkla\n(Ctrl ile birden fazla)",
            command=self.browse_files,
            fg_color="#3498DB",
            hover_color="#2980B9",
            text_color="white",
            font=("Segoe UI", 14, "bold"),
            height=80,
            corner_radius=12
        )
        self.drop_area.pack(fill="both", expand=True, padx=15, pady=15)

        # ── FILE LIST CARD ──
        file_list_card = self._create_card(content_frame, "📋 Seçili Dosyalar")
        file_list_card.grid(row=1, column=0, sticky="nsew", pady=(0, 20))

        tree_frame = ctk.CTkFrame(file_list_card, fg_color="#FFFFFF", corner_radius=8)
        tree_frame.pack(fill="both", expand=True, padx=15, pady=(10, 10))

        style = ttk.Style()
        style.theme_use('clam')
        style.configure('Treeview', rowheight=35, font=("Segoe UI", 11), background="#FFFFFF", foreground="#2C3E50", fieldbackground="#FFFFFF")
        style.configure('Treeview.Heading', font=("Segoe UI", 12, "bold"), background="#ECF0F1", foreground="#2C3E50")
        style.map('Treeview', background=[('selected', '#3498DB')], foreground=[('selected', 'white')])

        self.tree = ttk.Treeview(tree_frame, columns=("name", "items"), show="headings", height=8, selectmode="extended")
        self.tree.heading("name", text="📄 Dosya Adı")
        self.tree.heading("items", text="Durum")
        self.tree.column("name", anchor="w", width=300)
        self.tree.column("items", anchor="center", width=120)
        self.tree.tag_configure('even', background='#F8FBFF')
        self.tree.tag_configure('odd', background='#FFFFFF')

        scrollbar = ttk.Scrollbar(tree_frame, command=self.tree.yview)
        self.tree.configure(yscrollcommand=scrollbar.set)
        self.tree.pack(side="left", fill="both", expand=True)
        scrollbar.pack(side="right", fill="y")

        # BUTTONS
        btn_frame = ctk.CTkFrame(file_list_card, fg_color="#FFFFFF")
        btn_frame.pack(fill="x", padx=15, pady=(0, 15))

        btn_add = ctk.CTkButton(btn_frame, text="➕ Ekle", command=self.browse_files, fg_color="#27AE60", hover_color="#229954", text_color="white", font=("Segoe UI", 11, "bold"), width=100, corner_radius=8)
        btn_add.pack(side="left", padx=(0, 8))
        btn_del = ctk.CTkButton(btn_frame, text="🗑️ Sil", command=self.remove_selected, fg_color="#E74C3C", hover_color="#C0392B", text_color="white", font=("Segoe UI", 11, "bold"), width=100, corner_radius=8)
        btn_del.pack(side="left", padx=4)
        btn_clear = ctk.CTkButton(btn_frame, text="🔄 Temizle", command=self.clear_all, fg_color="#95A5A6", hover_color="#7F8C8D", text_color="white", font=("Segoe UI", 11, "bold"), width=100, corner_radius=8)
        btn_clear.pack(side="left", padx=4)

        # Ayırıcı
        ctk.CTkFrame(btn_frame, fg_color="#ECF0F1", width=2).pack(side="left", fill="y", padx=10, pady=2)

        # Sıralama butonları
        btn_up = ctk.CTkButton(btn_frame, text="🔼", command=self.move_up, fg_color="#8E44AD", hover_color="#7D3C98", text_color="white", font=("Segoe UI", 11, "bold"), width=50, corner_radius=8)
        btn_up.pack(side="left", padx=4)
        btn_down = ctk.CTkButton(btn_frame, text="🔽", command=self.move_down, fg_color="#8E44AD", hover_color="#7D3C98", text_color="white", font=("Segoe UI", 11, "bold"), width=50, corner_radius=8)
        btn_down.pack(side="left", padx=4)

        # Tooltip'ler
        Tooltip(btn_add, "Yeni dosya ekle")
        Tooltip(btn_del, "Seçili dosyaları sil")
        Tooltip(btn_clear, "Tüm listeyi temizle")
        Tooltip(btn_up, "Seçili dosyayı yukarı taşı")
        Tooltip(btn_down, "Seçili dosyayı aşağı taşı")

        # ── OUTPUT PATH CARD ──
        output_card = self._create_card(content_frame, "📁 Çıktı Konumu")
        output_card.grid(row=2, column=0, sticky="ew", pady=(0, 20))

        output_inner = ctk.CTkFrame(output_card, fg_color="#FFFFFF")
        output_inner.pack(fill="x", padx=15, pady=(10, 15))
        output_inner.grid_columnconfigure(0, weight=1)

        self.output_dir_label = ctk.CTkLabel(
            output_inner,
            text="İlk dosyanın klasörü (varsayılan)",
            font=("Segoe UI", 11),
            text_color="#7F8C8D",
            anchor="w"
        )
        self.output_dir_label.grid(row=0, column=0, sticky="ew", padx=(0, 10))

        btn_out = ctk.CTkButton(output_inner, text="📂 Seç", command=self.choose_output_dir, fg_color="#34495E", hover_color="#2C3E50", text_color="white", font=("Segoe UI", 11, "bold"), width=80, corner_radius=8)
        btn_out.grid(row=0, column=1)
        btn_reset = ctk.CTkButton(output_inner, text="↺", command=self.reset_output_dir, fg_color="#95A5A6", hover_color="#7F8C8D", text_color="white", font=("Segoe UI", 11, "bold"), width=40, corner_radius=8)
        btn_reset.grid(row=0, column=2, padx=(5, 0))
        Tooltip(btn_out, "Çıktı klasörü seç")
        Tooltip(btn_reset, "Varsayılana sıfırla")

        # ── STATUS CARD ──
        status_card = self._create_card(content_frame, "⚙️ Durum")
        status_card.grid(row=3, column=0, sticky="ew", pady=(0, 20))

        self.status_label = ctk.CTkLabel(status_card, text="✅ Hazır", font=("Segoe UI", 12), text_color="#27AE60")
        self.status_label.pack(anchor="w", padx=15, pady=(10, 5))

        self.progress = ctk.CTkProgressBar(status_card, fg_color="#ECF0F1", progress_color="#3498DB", height=6, corner_radius=3)
        self.progress.pack(fill="x", padx=15, pady=(5, 15))
        self.progress.set(0)

        # ── OPTIONS ──
        options_frame = ctk.CTkFrame(content_frame, fg_color="transparent")
        options_frame.grid(row=4, column=0, sticky="ew", pady=(0, 10))

        self.auto_open_var = ctk.BooleanVar(value=self._load_setting('auto_open', False))
        ctk.CTkCheckBox(
            options_frame,
            text="Bitince dosyayı otomatik aç",
            variable=self.auto_open_var,
            font=("Segoe UI", 12),
            text_color="#2C3E50",
            command=lambda: self._save_setting('auto_open', self.auto_open_var.get())
        ).pack(anchor="w")

        self.show_header_info_var = ctk.BooleanVar(value=self._load_setting('show_header_info', True))
        ctk.CTkCheckBox(
            options_frame,
            text="Sipariş bilgilerini göster (Tarih, RFQ, QTN)",
            variable=self.show_header_info_var,
            font=("Segoe UI", 12),
            text_color="#2C3E50",
            command=lambda: self._save_setting('show_header_info', self.show_header_info_var.get())
        ).pack(anchor="w", pady=(5, 0))

        # ── ACTION BUTTONS ──
        action_frame = ctk.CTkFrame(content_frame, fg_color="transparent")
        action_frame.grid(row=5, column=0, sticky="ew", pady=(10, 0))
        action_frame.grid_columnconfigure(0, weight=1)
        action_frame.grid_columnconfigure(1, weight=1)

        self.merge_btn = ctk.CTkButton(
            action_frame,
            text="🚀 Dosyaları Birleştir",
            command=self.merge_files,
            fg_color="#2980B9",
            hover_color="#1F618D",
            text_color="white",
            font=("Segoe UI", 14, "bold"),
            height=50,
            corner_radius=10,
            state="disabled"
        )
        self.merge_btn.grid(row=0, column=0, sticky="ew", padx=(0, 10))

        self.open_btn = ctk.CTkButton(
            action_frame,
            text="📄 Sonucu Aç",
            command=self.open_file,
            fg_color="#16A085",
            hover_color="#117A65",
            text_color="white",
            font=("Segoe UI", 14, "bold"),
            height=50,
            corner_radius=10,
            state="disabled"
        )
        self.open_btn.grid(row=0, column=1, sticky="ew")

        Tooltip(self.drop_area, "Excel dosyalarını seçmek için tıkla")
        Tooltip(self.merge_btn, "Seçili dosyaları tek bir Excel'de birleştir")
        Tooltip(self.open_btn, "Oluşturulan birleştirilmiş dosyayı aç")

        # İşlem sırasında kilitlenecek butonlar
        self._all_buttons = [
            self.drop_area, btn_add, btn_del, btn_clear,
            btn_up, btn_down, btn_out, btn_reset,
            self.merge_btn, self.open_btn
        ]

    def _create_card(self, parent, title):
        card = ctk.CTkFrame(parent, fg_color="#FFFFFF", corner_radius=12, border_width=1, border_color="#E8F4F8")
        ctk.CTkLabel(card, text=title, font=("Segoe UI", 13, "bold"), text_color="#2C3E50").pack(anchor="w", padx=15, pady=(15, 0))
        ctk.CTkFrame(card, fg_color="#ECF0F1", height=1).pack(fill="x", padx=15, pady=(10, 0))
        return card

    # ── Dosya İşlemleri ──────────────────────────────────────

    def browse_files(self):
        initial_dir = self._last_browse_dir if self._last_browse_dir else None
        files = filedialog.askopenfilenames(
            title="Excel Dosyalarını Seçin",
            initialdir=initial_dir,
            filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")]
        )
        if not files:
            return
        added = False
        for file_path in files:
            path = Path(file_path)
            if path.suffix.lower() == '.xlsx' and path not in self.uploaded_files:
                self.uploaded_files.append(path)
                added = True
        # Son klasörü kaydet
        self._last_browse_dir = str(Path(files[0]).parent)
        self._save_setting('last_browse_dir', self._last_browse_dir)
        if added:
            self._scan_and_update()

    def _scan_and_update(self):
        """Yeni eklenen dosyaları tara ve item sayısını göster"""
        self.update_file_list()
        threading.Thread(target=self._scan_worker, daemon=True).start()

    def _scan_worker(self):
        for f in list(self.uploaded_files):
            if f not in self.file_item_counts:
                data = self._extract_order_data(f)
                self.file_item_counts[f] = len(data['data_rows']) if data else -1
        self.root.after(0, self.update_file_list)

    def update_file_list(self):
        self.tree.delete(*self.tree.get_children())
        for i, f in enumerate(self.uploaded_files):
            count = self.file_item_counts.get(f)
            if count is None:
                status = "⏳ Taranıyor..."
            elif count < 0:
                status = "⚠️ Okunamadı"
            else:
                status = f"📊 {count} item"
            tag = 'even' if i % 2 == 0 else 'odd'
            self.tree.insert("", "end", values=(f.name, status), tags=(tag,))

        file_count = len(self.uploaded_files)
        if file_count > 0:
            total = sum(c for c in self.file_item_counts.values() if c and c > 0)
            self.status_label.configure(
                text=f"✅ {file_count} dosya seçildi ({total} item)" if total else f"✅ {file_count} dosya seçildi",
                text_color="#27AE60"
            )
            self.merge_btn.configure(state="normal")
        else:
            self.status_label.configure(text="⏳ Dosya seçin", text_color="#7F8C8D")
            self.merge_btn.configure(state="disabled")

    def remove_selected(self):
        """Seçili dosyaları sil (çoklu seçim destekli)"""
        selected = self.tree.selection()
        if not selected:
            return
        indices = sorted([self.tree.index(item) for item in selected], reverse=True)
        for idx in indices:
            if 0 <= idx < len(self.uploaded_files):
                removed = self.uploaded_files.pop(idx)
                self.file_item_counts.pop(removed, None)
        self.update_file_list()

    def clear_all(self):
        self.uploaded_files.clear()
        self.file_item_counts.clear()
        self.update_file_list()
        self.open_btn.configure(state="disabled")

    def move_up(self):
        selected = self.tree.selection()
        if not selected or len(selected) != 1:
            return
        idx = self.tree.index(selected[0])
        if idx > 0:
            self.uploaded_files[idx], self.uploaded_files[idx - 1] = self.uploaded_files[idx - 1], self.uploaded_files[idx]
            self.update_file_list()
            self.tree.selection_set(self.tree.get_children()[idx - 1])

    def move_down(self):
        selected = self.tree.selection()
        if not selected or len(selected) != 1:
            return
        idx = self.tree.index(selected[0])
        if idx < len(self.uploaded_files) - 1:
            self.uploaded_files[idx], self.uploaded_files[idx + 1] = self.uploaded_files[idx + 1], self.uploaded_files[idx]
            self.update_file_list()
            self.tree.selection_set(self.tree.get_children()[idx + 1])

    # ── Çıktı Konumu ─────────────────────────────────────────

    def choose_output_dir(self):
        dir_path = filedialog.askdirectory(title="Çıktı Klasörünü Seçin")
        if dir_path:
            self.custom_output_dir = Path(dir_path)
            self.output_dir_label.configure(text=str(self.custom_output_dir), text_color="#2C3E50")

    def reset_output_dir(self):
        self.custom_output_dir = None
        self.output_dir_label.configure(text="İlk dosyanın klasörü (varsayılan)", text_color="#7F8C8D")

    # ── Birleştirme ──────────────────────────────────────────

    def merge_files(self):
        if self.is_processing:
            return
        self.is_processing = True
        self._lock_ui()
        threading.Thread(target=self._merge_worker, daemon=True).start()

    def _check_write_permission(self, dir_path):
        """Klasöre yazma izni olup olmadığını kontrol et"""
        try:
            test_file = dir_path / '.write_test_tmp'
            test_file.touch()
            test_file.unlink()
            return True
        except Exception:
            return False

    def _is_file_locked(self, file_path):
        """Dosyanın başka bir program tarafından kilitli olup olmadığını kontrol et"""
        if not file_path.exists():
            return False
        try:
            with open(file_path, 'r+b'):
                return False
        except (IOError, PermissionError):
            return True

    def _merge_worker(self):
        try:
            self._update_progress(0)
            self._update_status("⏳ Şablon aranıyor...", "#F39C12")

            script_dir = _get_script_dir()
            self.template_path = script_dir / 'Final_List_Template.xlsx'

            if not self.template_path.exists():
                self._update_status("❌ Hata!", "#E74C3C")
                self.root.after(0, lambda: messagebox.showerror(
                    "Hata",
                    f"Template bulunamadı!\n\nLütfen Final_List_Template.xlsx dosyasını\nscript ile aynı klasöre koy.\n\n{script_dir}"
                ))
                return

            # Template erişim kontrolü
            if self._is_file_locked(self.template_path):
                self._update_status("❌ Hata!", "#E74C3C")
                self.root.after(0, lambda: messagebox.showerror(
                    "Hata",
                    "Template dosyası kilitli!\nExcel'de açıksa kapatıp tekrar deneyin."
                ))
                return

            # Çıktı klasörü yazma izni kontrolü
            output_dir = self.custom_output_dir or self.uploaded_files[0].parent
            if not self._check_write_permission(output_dir):
                self._update_status("❌ Hata!", "#E74C3C")
                err_dir = str(output_dir)
                self.root.after(0, lambda: messagebox.showerror(
                    "Hata",
                    f"Çıktı klasörüne yazılamıyor!\n{err_dir}"
                ))
                return

            self._update_progress(0.15)
            self._update_status("✅ Şablon bulundu", "#27AE60")
            time.sleep(0.5)

            timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
            self.output_path = output_dir / f'MERGED_FINAL_LIST_{timestamp}.xlsx'

            self._update_progress(0.3)
            self._update_status("📊 Dosyalar birleştiriliyor... (%30)", "#F39C12")
            self.root.after(0, self._start_pulse)

            total_items = self._create_merged_file()

            self.root.after(0, lambda: self._stop_pulse(0.8))
            self._update_status("🔄 Son işlemler... (%80)", "#F39C12")

            recalc_script = script_dir / 'recalc.py'
            if recalc_script.exists():
                try:
                    subprocess.run([sys.executable, str(recalc_script), str(self.output_path), '30'], capture_output=True, timeout=30)
                except Exception:
                    pass

            self._update_progress(1.0)
            file_count = len(self.uploaded_files)
            self._update_status(f"✅ Tamamlandı! ({file_count} sipariş, {total_items} item)", "#27AE60")

            # Otomatik aç veya bilgi göster
            if self.auto_open_var.get():
                self.root.after(0, self.open_file)
            else:
                out_name = self.output_path.name
                out_parent = str(self.output_path.parent)
                self.root.after(0, lambda: messagebox.showinfo(
                    "✅ Başarılı",
                    f"Final List oluşturuldu!\n\n📁 {out_name}\n📍 {out_parent}\n\n📊 {file_count} sipariş\n🔢 {total_items} item"
                ))

            # Doğrulama uyarısı
            self.root.after(300, self._show_verification_warning)

        except Exception as e:
            self.root.after(0, lambda: self._stop_pulse(0))
            self._update_status("❌ Hata!", "#E74C3C")
            error_msg = str(e)
            self.root.after(0, lambda: messagebox.showerror("Hata", f"Birleştirme hatası:\n{error_msg}"))
        finally:
            self.is_processing = False
            self.root.after(0, self._unlock_ui)

    def _show_verification_warning(self):
        """Birleştirme sonrası doğrulama uyarısı (3sn bekleme)"""
        dlg = ctk.CTkToplevel(self.root)
        dlg.title("⚠️ Önemli Uyarı")
        dlg.geometry("520x280")
        dlg.resizable(False, False)
        dlg.attributes('-topmost', True)
        dlg.grab_set()

        # Pencereyi ortala
        dlg.update_idletasks()
        x = self.root.winfo_x() + (self.root.winfo_width() - 520) // 2
        y = self.root.winfo_y() + (self.root.winfo_height() - 280) // 2
        dlg.geometry(f"+{x}+{y}")

        # Uyarı ikonu
        ctk.CTkLabel(
            dlg, text="⚠️", font=("Segoe UI", 48)
        ).pack(pady=(20, 5))

        # Uyarı metni
        ctk.CTkLabel(
            dlg,
            text="TOPLAM TUTARLARI MUTLAKA\nELLE KONTROL EDİNİZ!",
            font=("Segoe UI", 18, "bold"),
            text_color="#C0392B"
        ).pack(pady=(0, 5))

        ctk.CTkLabel(
            dlg,
            text="Birleştirilmiş dosyadaki tüm fiyat ve toplam değerlerini\ngöndermeden önce manuel olarak doğrulamanız gerekmektedir.",
            font=("Segoe UI", 11),
            text_color="#7F8C8D"
        ).pack(pady=(0, 15))

        # Buton (3sn geri sayım)
        btn = ctk.CTkButton(
            dlg,
            text="Anladım (3)",
            fg_color="#95A5A6",
            hover_color="#95A5A6",
            text_color="white",
            font=("Segoe UI", 13, "bold"),
            width=200, height=40,
            corner_radius=8,
            state="disabled"
        )
        btn.pack(pady=(0, 20))

        def countdown(sec):
            if sec > 0:
                btn.configure(text=f"Anladım ({sec})")
                dlg.after(1000, countdown, sec - 1)
            else:
                btn.configure(
                    text="Anladım ✓",
                    state="normal",
                    fg_color="#27AE60",
                    hover_color="#229954",
                    command=dlg.destroy
                )

        countdown(3)

    def _update_status(self, text, color):
        self.root.after(0, lambda: self.status_label.configure(text=text, text_color=color))

    def _update_progress(self, value):
        self.root.after(0, lambda: self.progress.set(value))

    def _start_pulse(self):
        """Progress bar'ı belirsiz (pulse) moduna al"""
        self._pulsing = True
        self._pulse_val = 0.0
        self._pulse_dir = 0.02
        self._do_pulse()

    def _do_pulse(self):
        if not self._pulsing:
            return
        self._pulse_val += self._pulse_dir
        if self._pulse_val >= 1.0 or self._pulse_val <= 0.0:
            self._pulse_dir *= -1
        self.progress.set(self._pulse_val)
        self.root.after(30, self._do_pulse)

    def _stop_pulse(self, final_value=1.0):
        """Pulse modunu durdur ve sabit değere ayarla"""
        self._pulsing = False
        self.progress.set(final_value)

    def _lock_ui(self):
        """İşlem sırasında tüm butonları kilitle"""
        for btn in self._all_buttons:
            btn.configure(state="disabled")

    def _unlock_ui(self):
        """İşlem sonrası butonları aç (merge butonu dosya durumuna göre)"""
        for btn in self._all_buttons:
            btn.configure(state="normal")
        if not self.uploaded_files:
            self.merge_btn.configure(state="disabled")
        if not (self.output_path and self.output_path.exists()):
            self.open_btn.configure(state="disabled")

    # ── Excel İşlemleri ──────────────────────────────────────

    def _create_merged_file(self):
        shutil.copy(self.template_path, self.output_path)
        wb = load_workbook(self.output_path)
        ws = wb.active

        template_start_row = 10
        for idx in range(1, 20):
            cell_value = ws.cell(idx, 1).value
            if cell_value and str(cell_value).strip().upper() == 'NO':
                template_start_row = idx + 1
                break

        current_row = template_start_row
        headers = ['NO', 'DESCRIPTION', 'CODE', 'QTTY', 'UNIT', 'U.PRICE', 'T.PRICE', 'REMARKS']
        total_items = 0

        for row in range(template_start_row, min(template_start_row + 1000, ws.max_row + 1)):
            for col in range(1, 12):
                cell = ws.cell(row, col)
                cell.border = self._no_border
                cell.number_format = 'General'
                cell.value = None

        # Her siparişin toplam satır referanslarını topla
        all_total_rows = []
        all_disc_rows = []
        all_gtotal_rows = []
        last_currency_symbol = ''

        for file_path in self.uploaded_files:
            order_data = self._extract_order_data(file_path)
            if not order_data:
                continue

            info = order_data['header_info']
            info_text = f"Order: {order_data['file_name']}"
            if 'rfq_ref' in info:
                info_text += f" | RFQ: {info['rfq_ref']}"
            if 'qtn_ref' in info:
                info_text += f" | QTN: {info['qtn_ref']}"

            currency = info.get('currency', '')
            currency_symbol = CURRENCY_SYMBOLS.get(currency.upper(), currency) if currency else ''
            if currency:
                info_text += f" | {currency}"

            # Sipariş bilgileri (A3:B5) sağ üst köşede + info text solda
            show_cells = self.show_header_info_var.get()
            header_cells = order_data.get('header_cells', [])
            has_cells = show_cells and any(l or v for l, v in header_cells)

            if has_cells:
                info_label_font = Font(bold=True, size=9)
                info_value_font = Font(size=9)
                # İlk satıra info text (sol) + ilk header cell (sağ)
                ws.cell(current_row, 2).value = info_text
                ws.cell(current_row, 2).font = Font(italic=True, size=9, color='808080')
                for i, (label, value) in enumerate(header_cells):
                    r = current_row + i
                    if label or value:
                        cell_g = ws.cell(r, 7)
                        clean_label = label.rstrip(' :')
                        cell_g.value = f"{clean_label} : " if clean_label else ''
                        cell_g.font = info_label_font
                        cell_g.alignment = self._right_align
                        cell_g.border = self._thin_border
                        cell_h = ws.cell(r, 8)
                        cell_h.value = value
                        cell_h.font = info_value_font
                        cell_h.border = self._thin_border
                current_row += len(header_cells)
            else:
                ws.cell(current_row, 2).value = info_text
                ws.cell(current_row, 2).font = Font(italic=True, size=9, color='808080')
                current_row += 1

            for col_idx, header in enumerate(headers, start=1):
                ws.cell(current_row, col_idx).value = header
            self._apply_header_style(ws, current_row)
            current_row += 1

            item_count = 0
            data_start_row = current_row
            price_format = f'"{currency_symbol}"#,##0.00' if currency_symbol else '#,##0.00'

            for data_row in order_data['data_rows']:
                item_count += 1
                for col_idx, value in enumerate(data_row, start=1):
                    cell = ws.cell(current_row, col_idx)
                    if col_idx == 1:
                        cell.value = item_count
                    elif col_idx == 7:
                        cell.value = f"=D{current_row}*F{current_row}"
                        cell.number_format = price_format
                    else:
                        cell.value = value
                        if col_idx == 6 and value is not None:
                            cell.number_format = price_format
                self._apply_data_row_style(ws, current_row)
                current_row += 1

            total_items += item_count
            current_row += 1

            self._apply_total_style(ws, current_row, 'TOTAL:')
            ws.cell(current_row, 7).value = f"=SUM(G{data_start_row}:G{data_start_row + item_count - 1})"
            ws.cell(current_row, 7).number_format = price_format
            total_row = current_row
            all_total_rows.append(current_row)
            current_row += 1

            disc_pct = info.get('discount_pct', 10)
            self._apply_total_style(ws, current_row, f'DISC.({disc_pct}%):')
            ws.cell(current_row, 7).value = f"=G{total_row}*{disc_pct/100}"
            ws.cell(current_row, 7).number_format = price_format
            disc_row = current_row
            all_disc_rows.append(current_row)
            current_row += 1

            self._apply_total_style(ws, current_row, 'G. TOTAL:')
            ws.cell(current_row, 7).value = f"=G{total_row}-G{disc_row}"
            ws.cell(current_row, 7).number_format = price_format
            all_gtotal_rows.append(current_row)
            last_currency_symbol = currency_symbol
            current_row += 4

        # ── GRAND SUMMARY ──
        if len(all_gtotal_rows) > 1:
            summary_format = f'"{last_currency_symbol}"#,##0.00' if last_currency_symbol else '#,##0.00'

            # Ayırıcı çizgi
            separator_fill = PatternFill(start_color='2C3E50', end_color='2C3E50', fill_type='solid')
            for col in range(1, 9):
                cell = ws.cell(current_row, col)
                cell.fill = separator_fill
                cell.border = self._thin_border
            current_row += 1

            # Başlık satırı
            banner_fill = PatternFill(start_color='1A5276', end_color='1A5276', fill_type='solid')
            banner_font = Font(bold=True, size=13, color='FFFFFF')
            ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=8)
            title_cell = ws.cell(current_row, 1)
            title_cell.value = f'GRAND SUMMARY  —  {len(all_gtotal_rows)} ORDERS'
            title_cell.font = banner_font
            title_cell.fill = banner_fill
            title_cell.alignment = Alignment(horizontal='center', vertical='center')
            title_cell.border = self._thin_border
            for col in range(2, 9):
                ws.cell(current_row, col).fill = banner_fill
                ws.cell(current_row, col).border = self._thin_border
            current_row += 1

            # Boş ayırıcı
            current_row += 1

            # TOTAL formülü
            total_refs = '+'.join([f'G{r}' for r in all_total_rows])
            summary_label_font = Font(bold=True, size=12, color='2C3E50')
            summary_value_font = Font(bold=True, size=12, color='1A5276')
            summary_border = Border(
                left=Side(style='medium'), right=Side(style='medium'),
                top=Side(style='medium'), bottom=Side(style='medium')
            )
            label_fill = PatternFill(start_color='EBF5FB', end_color='EBF5FB', fill_type='solid')
            value_fill = PatternFill(start_color='D4E6F1', end_color='D4E6F1', fill_type='solid')

            # TOTAL
            ws.merge_cells(start_row=current_row, start_column=4, end_row=current_row, end_column=6)
            lbl = ws.cell(current_row, 4)
            lbl.value = 'TOTAL :'
            lbl.font = summary_label_font
            lbl.alignment = Alignment(horizontal='right', vertical='center')
            lbl.fill = label_fill
            lbl.border = summary_border
            for c in range(5, 7):
                ws.cell(current_row, c).fill = label_fill
                ws.cell(current_row, c).border = summary_border
            ws.merge_cells(start_row=current_row, start_column=7, end_row=current_row, end_column=8)
            val = ws.cell(current_row, 7)
            val.value = f'={total_refs}'
            val.font = summary_value_font
            val.number_format = summary_format
            val.alignment = Alignment(horizontal='center', vertical='center')
            val.fill = value_fill
            val.border = summary_border
            ws.cell(current_row, 8).fill = value_fill
            ws.cell(current_row, 8).border = summary_border
            current_row += 1

            # DISCOUNT
            disc_refs = '+'.join([f'G{r}' for r in all_disc_rows])
            ws.merge_cells(start_row=current_row, start_column=4, end_row=current_row, end_column=6)
            lbl = ws.cell(current_row, 4)
            lbl.value = 'TOTAL DISCOUNT :'
            lbl.font = summary_label_font
            lbl.alignment = Alignment(horizontal='right', vertical='center')
            lbl.fill = label_fill
            lbl.border = summary_border
            for c in range(5, 7):
                ws.cell(current_row, c).fill = label_fill
                ws.cell(current_row, c).border = summary_border
            ws.merge_cells(start_row=current_row, start_column=7, end_row=current_row, end_column=8)
            val = ws.cell(current_row, 7)
            val.value = f'={disc_refs}'
            val.font = summary_value_font
            val.number_format = summary_format
            val.alignment = Alignment(horizontal='center', vertical='center')
            val.fill = value_fill
            val.border = summary_border
            ws.cell(current_row, 8).fill = value_fill
            ws.cell(current_row, 8).border = summary_border
            current_row += 1

            # GRAND TOTAL
            gtotal_refs = '+'.join([f'G{r}' for r in all_gtotal_rows])
            grand_fill = PatternFill(start_color='1A5276', end_color='1A5276', fill_type='solid')
            grand_font = Font(bold=True, size=14, color='FFFFFF')
            ws.merge_cells(start_row=current_row, start_column=4, end_row=current_row, end_column=6)
            lbl = ws.cell(current_row, 4)
            lbl.value = 'GRAND TOTAL :'
            lbl.font = grand_font
            lbl.alignment = Alignment(horizontal='right', vertical='center')
            lbl.fill = grand_fill
            lbl.border = summary_border
            for c in range(5, 7):
                ws.cell(current_row, c).fill = grand_fill
                ws.cell(current_row, c).border = summary_border
            ws.merge_cells(start_row=current_row, start_column=7, end_row=current_row, end_column=8)
            val = ws.cell(current_row, 7)
            val.value = f'={gtotal_refs}'
            val.font = grand_font
            val.number_format = summary_format
            val.alignment = Alignment(horizontal='center', vertical='center')
            val.fill = grand_fill
            val.border = summary_border
            ws.cell(current_row, 8).fill = grand_fill
            ws.cell(current_row, 8).border = summary_border
            current_row += 2

        ws.column_dimensions['A'].width = 8
        ws.column_dimensions['B'].width = 65
        ws.column_dimensions['C'].width = 15
        ws.column_dimensions['D'].width = 10
        ws.column_dimensions['E'].width = 10
        ws.column_dimensions['F'].width = 12
        ws.column_dimensions['G'].width = 12
        ws.column_dimensions['H'].width = 30

        last_row = current_row - 1
        ws.print_area = f'A1:H{last_row}'
        ws.sheet_view.showGridLines = False

        wb.save(self.output_path)
        return total_items

    def _extract_order_data(self, file_path):
        try:
            df = pd.read_excel(file_path, header=None)

            header_info = {}
            if len(df.columns) < 2:
                return None
            for idx in range(min(15, len(df))):
                first_col = str(df.iloc[idx, 0]).strip() if pd.notna(df.iloc[idx, 0]) else ''
                second_col = df.iloc[idx, 1] if pd.notna(df.iloc[idx, 1]) else ''

                if 'RFQ REF' in first_col.upper():
                    header_info['rfq_ref'] = second_col
                elif 'QTN REF' in first_col.upper():
                    header_info['qtn_ref'] = second_col
                elif 'CURRENCY' in first_col.upper():
                    header_info['currency'] = str(second_col).strip()
                elif 'DISC' in first_col.upper() and '%' in first_col.upper():
                    try:
                        header_info['discount_pct'] = float(second_col)
                    except (ValueError, TypeError):
                        header_info['discount_pct'] = 10

            if 'discount_pct' not in header_info:
                header_info['discount_pct'] = 10

            # A3:B5 hücrelerini çıkar (Tarih, RFQ REF, QTN REF)
            header_cells = []
            for row_idx in range(2, 5):  # Excel satır 3,4,5 -> 0-indexed 2,3,4
                if row_idx < len(df):
                    label = str(df.iloc[row_idx, 0]).strip() if pd.notna(df.iloc[row_idx, 0]) else ''
                    value = str(df.iloc[row_idx, 1]).strip() if pd.notna(df.iloc[row_idx, 1]) else ''
                    header_cells.append((label, value))
                else:
                    header_cells.append(('', ''))

            start_row = None
            for idx in range(len(df)):
                if df.iloc[idx, 0] == 'NO':
                    start_row = idx
                    break

            if start_row is None:
                return None

            data_rows = []
            for idx in range(start_row + 1, len(df)):
                row = df.iloc[idx]
                first_col_val = row.iloc[0]
                # TOTAL satırı: A sütunu boş/NaN ve satırda TOTAL geçiyor
                if pd.isna(first_col_val) or str(first_col_val).strip() == '':
                    row_str = ' '.join([str(x).upper() for x in row.values if pd.notna(x)])
                    if 'TOTAL' in row_str:
                        break
                    continue

                # Sıra numarası kontrolü: 1, 2, 3 veya 1A, 1B, 2A gibi
                val_str = str(first_col_val).strip()
                if val_str and val_str[0].isdigit():
                    data_rows.append(row.values.tolist())
                else:
                    continue

            return {
                'file_name': Path(file_path).name,
                'header_info': header_info,
                'header_cells': header_cells,
                'data_rows': data_rows,
            }
        except Exception:
            return None

    # ── Stiller ──────────────────────────────────────────────

    def _apply_header_style(self, ws, row_num):
        for col in range(1, 9):
            cell = ws.cell(row_num, col)
            cell.fill = self._header_fill
            cell.font = self._header_font
            cell.alignment = self._center_align
            cell.border = self._thin_border

    def _apply_data_row_style(self, ws, row_num):
        for col in range(1, 9):
            cell = ws.cell(row_num, col)
            cell.border = self._thin_border
            cell.alignment = self._data_align

    def _apply_total_style(self, ws, row_num, label):
        for col in range(1, 9):
            ws.cell(row_num, col).border = self._no_border
        ws.cell(row_num, 6).value = label
        ws.cell(row_num, 6).font = self._bold_font
        ws.cell(row_num, 6).alignment = self._right_align
        ws.cell(row_num, 7).font = self._bold_font

    # ── Dosya Açma ───────────────────────────────────────────

    def open_file(self):
        if self.output_path and self.output_path.exists():
            try:
                os.startfile(str(self.output_path))
            except Exception:
                pass


def main():
    if HAS_DND:
        class DnDCTk(ctk.CTk, TkinterDnD.DnDWrapper):
            def __init__(self):
                super().__init__()
                self.TkdndVersion = TkinterDnD._require(self)
        root = DnDCTk()
    else:
        root = ctk.CTk()
    FinalListMerger(root)
    root.mainloop()


if __name__ == '__main__':
    main()
